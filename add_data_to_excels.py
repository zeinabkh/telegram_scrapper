import pandas as pd
from telethon import TelegramClient
import datetime
from pytz import timezone
import xlsxwriter as xlwr
from openpyxl import workbook
from openpyxl import load_workbook
from telethon.tl.functions.messages import GetHistoryRequest
import openpyxl as pyxl
# Remember to use your own values from my.telegram.org!

api_id = 1790282
api_hash = 'ca3cd83dbafd959395e408e9575c31e0'
client = TelegramClient('anon', api_id, api_hash)
def message_process(new_message):
    post= {}
    post['post_id'] = int(str(id) + str(new_message.id))
    post['message'] = new_message.text
    post['sender_username'] = new_message.username
    # post['sender_title'] = utils.get_display_name(sender)
    print(0)
    try:
        if new_message.views:
            post['view'] = new_message.views
        else:
            post['view'] = -1
    except KeyError:
        post['view'] = -1
    post['process'] = 'false'
    post['date_get'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    post['date_published'] = new_message.date.astimezone(timezone('Asia/Tehran')).strftime("%Y-%m-%d %H:%M:%S")
    print(post)


async def main():


    group_list = [{'name':"سهامداران شتران شپنا",'date':datetime.datetime.now()}]
    for group in group_list:
        data_load = pyxl.load_workbook(group['name'] + ".xlsx")
        last_message_id = data_load.get_sheet_by_name(data_load.sheetnames[1]).cell(2,5).value
        print(last_message_id)
        previous_date = (group['date'].year, group['date'].month, group['date'].day)
        date_sheet = data_load.create_sheet(str(group['date'].year)+"-" + str(group['date'].month)+"-" +str(group['date'].day),index=0)
        date_sheet.cell(1,2).value = 'user_id'
        date_sheet.cell(1, 3).value =   "user_name"
        date_sheet.cell(1, 4).value =  "name"
        date_sheet.cell(1, 5).value =  "message_id"
        date_sheet.cell(1,6).value = "text"
        i = 2
        j = 1
        async for message in client.iter_messages(group['name'], offset_date=group['date'], min_id=last_message_id):

            j += 1
            message_date = (message.date.year, message.date.month, message.date.day)
            if message_date == previous_date:
                if message.media is None and message.from_id is not None:
                    user = await client.get_entity(message.from_id.user_id)
                    if user.last_name is not None:
                        name = user.first_name + ' ' + user.last_name
                    else:
                        name = user.first_name
                    date_sheet.cell(i, 1).value = message.from_id.user_id
                    date_sheet.cell(i, 2).value = user.username
                    date_sheet.cell(i, 3).value = name
                    date_sheet.cell(i, 4).value = message.id
                    date_sheet.cell(i, 5).value = message.text
                    print("write2")
                    i += 1

            else:
                i = 2
                previous_date = (message.date.year, message.date.month, message.date.day)
                date_sheet = data_load.create_sheet(
                    str(message.date.year) + "-" + str( message.date.month) + "-" + str(message.date.day),index= 0)
                date_sheet.cell(1, 2).value = 'user_id'
                date_sheet.cell(1, 3).value = "user_name"
                date_sheet.cell(1, 4).value = "name"
                date_sheet.cell(1, 5).value = "message_id"
                date_sheet.cell(1, 6).value = "text"
                if message.media is None and message.from_id is not None:
                    user = await client.get_entity(message.from_id.user_id)
                    if user.last_name is not None:
                        name = user.first_name + ' ' + user.last_name
                    else:
                        name = user.first_name
                    date_sheet.cell(i, 1).value = message.from_id.user_id
                    date_sheet.cell(i, 2).value = user.username
                    date_sheet.cell(i, 3).value = name
                    date_sheet.cell(i, 4).value = message.id
                    date_sheet.cell(i, 5).value = message.text
                    i += 1
                print("write1")
        data_load.save(group['name'] + ".xlsx")
with client:
    client.loop.run_until_complete(main())